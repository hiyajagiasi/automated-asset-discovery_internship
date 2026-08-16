from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _apply_sheet_formatting(sheet, header_fill: PatternFill | None = None) -> None:
    sheet.freeze_panes = "A2"
    if sheet.max_row > 1:
        sheet.auto_filter.ref = sheet.dimensions

    for row_idx in range(1, sheet.max_row + 1):
        for col_idx in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if row_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill or PatternFill("solid", fgColor="1F4E78")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
                if isinstance(cell.value, str) and cell.value.startswith(("http://", "https://")):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"

    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column
        for cell in column_cells:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 70)
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width


def generate_excel_report(
    base_dir: Path,
    target: str,
    subdomains: list[str],
    live_hosts: list[str],
    *args,
    config: dict[str, Any] | None = None,
    dmarc: list[dict[str, str]] | None = None,
) -> Path:
    dead_hosts: list[str] = []
    ports: list[dict[str, str]] = []
    technologies: list[dict[str, str]] = []

    if args:
        if len(args) >= 1 and isinstance(args[-1], dict):
            config = args[-1]
            args = args[:-1]

        if len(args) == 3:
            dead_hosts, ports, technologies = args
        elif len(args) == 2:
            ports, technologies = args
        elif len(args) == 1:
            first = args[0]
            if isinstance(first, list) and first and isinstance(first[0], dict):
                ports = first
            else:
                dead_hosts = first

    if config is None:
        config = {}

    if dmarc is None:
        dmarc = []

    report_cfg = config.get("reports", {}) if isinstance(config, dict) else {}
    output_path = Path(report_cfg.get("excel", "reports/report.xlsx"))
    output_path.parent.mkdir(parents=True, exist_ok=True)

    generated_at = pd.Timestamp.now(tz="UTC").strftime("%Y-%m-%d %H:%M:%S UTC")
    summary_rows = [
        ["Total Subdomains", len(subdomains)],
        ["Live Hosts", len(live_hosts)],
        ["Dead Hosts", len(dead_hosts)],
        ["Open Ports", len(ports)],
        ["Technologies Detected", len(technologies)],
        ["DMARC Hosts", sum(1 for item in dmarc if item.get("status") in {"configured", "inherited"})],
        ["DMARC Missing", sum(1 for item in dmarc if item.get("status") == "missing")],
        ["Scan Date", datetime.now().strftime("%d-%m-%Y")],
    ]
    summary = pd.DataFrame(summary_rows, columns=["Metric", "Count"])

    normalized_ports = []
    for item in ports:
        if isinstance(item, dict):
            normalized_ports.append({
                "host": item.get("host", "unknown"),
                "port": item.get("port", "unknown"),
                "service": item.get("service", "unknown"),
            })

    normalized_technologies = []
    for item in technologies:
        if isinstance(item, dict):
            normalized_technologies.append({
                "host": item.get("host", "unknown"),
                "technology": item.get("technology", "unknown"),
            })

    normalized_dmarc = []
    for item in dmarc:
        if isinstance(item, dict):
            normalized_dmarc.append({
                "Host": item.get("host", "unknown"),
                "Status": (item.get("status") or "unknown").upper(),
                "Policy": (item.get("policy") or "none").upper(),
                "Policy Type": "MONITORING" if (item.get("policy") or "none").lower() == "none" else "ENFORCEMENT" if (item.get("policy") or "none").lower() in {"quarantine", "reject"} else "UNKNOWN",
                "Source": item.get("source_domain") or item.get("source") or "NONE",
                "Source Type": (item.get("source_type") or "none").upper(),
                "Policy Source": (
                    f"{(item.get('source_type') or 'none').upper()} {(item.get('policy_source') or 'none').upper()}"
                    if (item.get("source_type") or "none") not in {"none", ""} and (item.get("policy_source") or "none") not in {"none", ""}
                    else "NONE"
                ),
                "Record": item.get("dmarc_record") or "NONE",
            })

    subdomain_df = pd.DataFrame({"Sr No": range(1, len(subdomains) + 1), "Subdomain": subdomains}) if subdomains else pd.DataFrame(columns=["Sr No", "Subdomain"])
    live_host_df = pd.DataFrame({"Host": live_hosts, "Status Code": "", "IP": "", "Title": ""}) if live_hosts else pd.DataFrame(columns=["Host", "Status Code", "IP", "Title"])
    port_df = pd.DataFrame(normalized_ports) if normalized_ports else pd.DataFrame(columns=["Host", "Port", "Protocol", "Service"])
    if not normalized_ports:
        port_df = pd.DataFrame(columns=["Host", "Port", "Protocol", "Service"])
    else:
        port_df = pd.DataFrame([
            {
                "Host": item.get("host", "unknown"),
                "Port": item.get("port", "unknown"),
                "Protocol": "TCP",
                "Service": item.get("service", "unknown"),
            }
            for item in normalized_ports
        ])

    technology_df = pd.DataFrame(
        [
            {"Host": item.get("host", "unknown"), "Technology": item.get("technology", "unknown").split("|")[0].strip(), "Version": ""}
            for item in normalized_technologies
        ]
    ) if normalized_technologies else pd.DataFrame(columns=["Host", "Technology", "Version"])

    dmarc_df = pd.DataFrame(normalized_dmarc) if normalized_dmarc else pd.DataFrame(columns=["Host", "Status", "Policy", "Policy Type", "Source", "Source Type", "Policy Source", "Record"])

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        subdomain_df.to_excel(writer, sheet_name="Subdomains", index=False)
        live_host_df.to_excel(writer, sheet_name="Live Hosts", index=False)
        port_df.to_excel(writer, sheet_name="Open Ports", index=False)
        technology_df.to_excel(writer, sheet_name="Technologies", index=False)
        dmarc_df.to_excel(writer, sheet_name="DMARC", index=False)

    workbook = load_workbook(output_path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        _apply_sheet_formatting(sheet, header_fill=header_fill)

    workbook.save(output_path)
    return output_path
